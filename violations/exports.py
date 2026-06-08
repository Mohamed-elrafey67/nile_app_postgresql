"""
أدوات تصدير التقارير — Excel و PDF + تقارير الأقمار الصناعية
"""
import io
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                 Paragraph, Spacer, HRFlowable, Image)
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.enums import TA_RIGHT, TA_CENTER, TA_LEFT
import os

# ── Arabic text RTL reordering for ReportLab ───────────────────────────
import re as _re
def ar(text):
    """Reshape + reorder Arabic text for visual display in ReportLab.
    Returns visual-order text (LTR) — use TA_LEFT for alignment.
    Strips <b> tags (bold not needed — use separate styles)."""
    try:
        from bidi.algorithm import get_display
        import arabic_reshaper
        clean = _re.sub(r'</?b>', '', text)
        reshaped = arabic_reshaper.reshape(clean)
        return get_display(reshaped)
    except ImportError:
        return text

# ── COLUMN DEFINITIONS ─────────────────────────────────────────────────────
COLUMNS = [
    ('code',               'الرمز',                    12),
    ('governorate__name_ar','المحافظة',                 14),
    ('district_name',      'المركز',                   14),
    ('village',            'القرية / المدينة',          18),
    ('occupant',           'اسم المستغل',              20),
    ('basin',              'اسم الحوض',                22),
    ('description',        'وصف الاستغلال',            18),
    ('area_nile_bank',     'مساحة على الجسر م²',        14),
    ('area_public',        'مساحة على المنفعة م²',      14),
    ('area_outside',       'خارج الحياض م²',           13),
    ('area_total',         'المسطح الإجمالي م²',       16),
    ('status',             'الحالة',                    10),
]

STATUS_AR = {'approved': 'معتمد', 'pending': 'في الانتظار', 'rejected': 'مرفوض'}


# ══════════════════════════════════════════════════════════════════
# EXCEL EXPORT
# ══════════════════════════════════════════════════════════════════
def export_excel(records, filters=None, user=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'تواجدات طرح النهر'
    ws.sheet_view.rightToLeft = True

    # ── Colors ──
    HDR_FILL  = PatternFill('solid', fgColor='0D3B6E')
    SUB_FILL  = PatternFill('solid', fgColor='1A6EA8')
    ALT_FILL  = PatternFill('solid', fgColor='EEF4FB')
    GRN_FILL  = PatternFill('solid', fgColor='D4EDDA')
    YEL_FILL  = PatternFill('solid', fgColor='FFF3CD')
    RED_FILL  = PatternFill('solid', fgColor='F8D7DA')
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Title Row ──
    ws.merge_cells('A1:L1')
    title_cell = ws['A1']
    title_cell.value = 'منظومة توثيق أراضي طرح النهر — وزارة الموارد المائية والري'
    title_cell.font  = Font(name='Arial', bold=True, size=14, color='FFFFFF')
    title_cell.fill  = HDR_FILL
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # ── Subtitle ──
    ws.merge_cells('A2:L2')
    sub = ws['A2']
    now = datetime.now().strftime('%Y-%m-%d %H:%M')
    filter_text = _build_filter_text(filters)
    sub.value = f'تاريخ التقرير: {now}  |  {filter_text}  |  إجمالي السجلات: {len(records)}'
    sub.font  = Font(name='Arial', size=10, color='FFFFFF')
    sub.fill  = SUB_FILL
    sub.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 20

    # ── Headers ──
    ws.row_dimensions[3].height = 22
    for col_idx, (field, label, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=3, column=col_idx, value=label)
        cell.font      = Font(name='Arial', bold=True, size=10, color='FFFFFF')
        cell.fill      = PatternFill('solid', fgColor='1A5276')
        cell.alignment = Alignment(horizontal='center', vertical='center',
                                   wrap_text=True)
        cell.border    = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Data Rows ──
    total_area = 0
    for row_idx, rec in enumerate(records, 4):
        fill = ALT_FILL if row_idx % 2 == 0 else PatternFill('solid', fgColor='FFFFFF')
        ws.row_dimensions[row_idx].height = 16

        for col_idx, (field, label, width) in enumerate(COLUMNS, 1):
            val = rec.get(field, '')
            if field == 'status':
                val = STATUS_AR.get(str(val), val)
                if rec.get('status') == 'approved':   fill = GRN_FILL if row_idx % 2 == 0 else PatternFill('solid', fgColor='ECFDF5')
                elif rec.get('status') == 'pending':  fill = YEL_FILL if row_idx % 2 == 0 else PatternFill('solid', fgColor='FFFDE7')
                elif rec.get('status') == 'rejected': fill = RED_FILL if row_idx % 2 == 0 else PatternFill('solid', fgColor='FFF5F5')
            if field in ('area_nile_bank','area_public','area_outside','area_total'):
                val = round(float(val or 0), 2)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font      = Font(name='Arial', size=9)
            cell.alignment = Alignment(horizontal='center' if field.startswith('area') or field == 'code' or field == 'status'
                                       else 'right', vertical='center')
            cell.fill   = fill
            cell.border = border
        total_area += float(rec.get('area_total', 0) or 0)

    # ── Totals Row ──
    total_row = len(records) + 4
    ws.row_dimensions[total_row].height = 20
    ws.merge_cells(f'A{total_row}:J{total_row}')
    tot_label = ws[f'A{total_row}']
    tot_label.value = f'الإجمالي — {len(records)} سجل'
    tot_label.font  = Font(name='Arial', bold=True, size=10, color='FFFFFF')
    tot_label.fill  = HDR_FILL
    tot_label.alignment = Alignment(horizontal='center', vertical='center')
    tot_label.border = border

    tot_val = ws[f'K{total_row}']
    tot_val.value = round(total_area, 2)
    tot_val.font  = Font(name='Arial', bold=True, size=10, color='FFFFFF')
    tot_val.fill  = HDR_FILL
    tot_val.alignment = Alignment(horizontal='center', vertical='center')
    tot_val.border = border

    ws[f'L{total_row}'].fill = HDR_FILL
    ws[f'L{total_row}'].border = border

    # ── Summary Sheet ──
    ws2 = wb.create_sheet('ملخص إحصائي')
    ws2.sheet_view.rightToLeft = True
    _build_summary_sheet(ws2, records, HDR_FILL, SUB_FILL, border)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _build_summary_sheet(ws, records, hdr_fill, sub_fill, border):
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 18

    def hdr(row, text):
        ws.merge_cells(f'A{row}:B{row}')
        c = ws[f'A{row}']
        c.value = text
        c.font  = Font(name='Arial', bold=True, size=11, color='FFFFFF')
        c.fill  = hdr_fill
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row].height = 22

    def row2(ws, row, label, val):
        lc = ws.cell(row=row, column=1, value=label)
        vc = ws.cell(row=row, column=2, value=val)
        lc.font = Font(name='Arial', size=10)
        vc.font = Font(name='Arial', bold=True, size=10)
        lc.alignment = Alignment(horizontal='right')
        vc.alignment = Alignment(horizontal='center')
        lc.border = vc.border = border

    total = len(records)
    total_area = sum(float(r.get('area_total', 0) or 0) for r in records)

    hdr(1, 'ملخص إحصائي للتقرير')
    row2(ws, 2, 'إجمالي السجلات', total)
    row2(ws, 3, 'إجمالي المساحة م²', round(total_area, 2))
    row2(ws, 4, 'متوسط المساحة م²', round(total_area / total, 2) if total else 0)

    # By status
    hdr(6, 'توزيع الحالات')
    from collections import Counter
    statuses = Counter(r.get('status', '') for r in records)
    r = 7
    for s, cnt in statuses.items():
        row2(ws, r, STATUS_AR.get(s, s), cnt)
        r += 1

    # By gov
    hdr(r + 1, 'توزيع المحافظات')
    govs = Counter(r.get('governorate__name_ar', '—') for r in records)
    r += 2
    for g, cnt in govs.most_common():
        row2(ws, r, g, cnt)
        r += 1

    # By description
    hdr(r + 1, 'أنواع الاستغلال (أعلى 10)')
    descs = Counter(r.get('description', '') for r in records)
    r += 2
    for d, cnt in descs.most_common(10):
        row2(ws, r, d[:40], cnt)
        r += 1


def _build_filter_text(filters):
    if not filters:
        return 'جميع البيانات'
    parts = []
    mapping = {
        'gov': 'المحافظة', 'district': 'المركز', 'village': 'القرية',
        'description': 'وصف الاستغلال', 'min_area': 'الحد الأدنى للمساحة',
        'search': 'بحث', 'status': 'الحالة',
    }
    for k, v in filters.items():
        if v:
            parts.append(f"{mapping.get(k, k)}: {v}")
    return ' | '.join(parts) if parts else 'جميع البيانات'


# ══════════════════════════════════════════════════════════════════
# PDF EXPORT
# ══════════════════════════════════════════════════════════════════
def export_pdf(records, filters=None, user=None):
    buf = io.BytesIO()

    # Register Arabic font if available
    font_name = 'Helvetica'  # fallback
    win_fonts = [
        'C:\\Windows\\Fonts\\arial.ttf',
        'C:\\Windows\\Fonts\\tahoma.ttf',
        'C:\\Windows\\Fonts\\times.ttf',
        'C:\\Windows\\Fonts\\Candarab.ttf',
    ]
    linux_fonts = [
        '/usr/share/fonts/truetype/arabic/ae_AlArabiya.ttf',
        '/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf',
    ]
    for fp in win_fonts + linux_fonts:
        if os.path.exists(fp):
            try:
                pdfmetrics.registerFont(TTFont('Arabic', fp))
                font_name = 'Arabic'
                break
            except Exception:
                pass

    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        rightMargin=1.5*cm, leftMargin=1.5*cm,
        topMargin=2*cm, bottomMargin=2*cm,
        title='تقرير تواجدات طرح النهر',
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'ArabicTitle', fontName=font_name, fontSize=14,
        textColor=colors.HexColor('#0D3B6E'),
        alignment=TA_CENTER, spaceAfter=6,
    )
    subtitle_style = ParagraphStyle(
        'ArabicSub', fontName=font_name, fontSize=9,
        textColor=colors.grey, alignment=TA_CENTER, spaceAfter=12,
    )
    normal_style = ParagraphStyle(
        'ArabicNormal', fontName=font_name, fontSize=8,
        alignment=TA_LEFT,
    )

    elements = []

    # Title
    elements.append(Paragraph(ar('منظومة توثيق أراضي طرح النهر'), title_style))
    elements.append(Paragraph(ar('وزارة الموارد المائية والري — جمهورية مصر العربية'), subtitle_style))
    elements.append(HRFlowable(width='100%', thickness=2,
                                color=colors.HexColor('#0D3B6E'), spaceAfter=8))

    # Meta info
    now = datetime.now().strftime('%Y-%m-%d %H:%M')
    filter_text = _build_filter_text(filters)
    meta = ar(f'تاريخ: {now}   |   {filter_text}   |   عدد السجلات: {len(records)}')
    elements.append(Paragraph(meta, subtitle_style))
    elements.append(Spacer(1, 0.3*cm))

    # Table data (reversed for RTL layout)
    headers = [ar(col[1]) for col in COLUMNS][::-1]
    data = [headers]
    total_area = 0
    for rec in records:
        row = []
        for field, label, width in COLUMNS:
            val = rec.get(field, '') or ''
            if field == 'status':
                val = STATUS_AR.get(str(val), val)
            elif field in ('area_nile_bank','area_public','area_outside','area_total'):
                val = f"{float(val or 0):.1f}"
            row.append(ar(str(val)[:30]))
        data.append(row[::-1])
        total_area += float(rec.get('area_total', 0) or 0)

    # Totals row (reversed for RTL)
    totals = [ar('الإجمالي')] + [''] * 9 + [f'{total_area:.1f}', '']
    data.append(totals[::-1])

    # Column widths proportional (reversed for RTL)
    col_widths = [
        1.5*cm, 2.2*cm, 2.2*cm, 2.8*cm, 3.2*cm,
        3.5*cm, 2.8*cm, 2.0*cm, 2.0*cm, 2.0*cm,
        2.2*cm, 1.6*cm,
    ][::-1]

    table = Table(data, colWidths=col_widths, repeatRows=1)

    # Style
    primary   = colors.HexColor('#0D3B6E')
    secondary = colors.HexColor('#1A6EA8')
    alt_row   = colors.HexColor('#EEF4FB')

    style = TableStyle([
        # Header
        ('BACKGROUND',  (0,0), (-1,0),  primary),
        ('TEXTCOLOR',   (0,0), (-1,0),  colors.white),
        ('FONTNAME',    (0,0), (-1,0),  font_name),
        ('FONTSIZE',    (0,0), (-1,0),  7.5),
        ('ALIGN',       (0,0), (-1,0),  'CENTER'),
        ('VALIGN',      (0,0), (-1,-1), 'MIDDLE'),
        ('ROWBACKGROUND',(0,1),(-1,-2), [colors.white, alt_row]),
        ('FONTNAME',    (0,1), (-1,-1), font_name),
        ('FONTSIZE',    (0,1), (-1,-1), 7),
        ('ALIGN',       (0,1), (-1,-1), 'CENTER'),
        ('GRID',        (0,0), (-1,-1), 0.4, colors.HexColor('#CCCCCC')),
        ('ROWHEIGHT',   (0,0), (-1,-1), 14),
        # Totals row
        ('BACKGROUND',  (0,-1), (-1,-1), primary),
        ('TEXTCOLOR',   (0,-1), (-1,-1), colors.white),
        ('FONTNAME',    (0,-1), (-1,-1), font_name),
        ('FONTSIZE',    (0,-1), (-1,-1), 8),
        ('FONTWEIGHT',  (0,-1), (-1,-1), 'BOLD'),
    ])
    table.setStyle(style)
    elements.append(table)

    # Footer summary
    elements.append(Spacer(1, 0.5*cm))
    elements.append(HRFlowable(width='100%', thickness=1,
                                color=colors.HexColor('#CCCCCC'), spaceAfter=6))
    summary_text = ar(f'إجمالي المساحة المتعدى عليها: {total_area:,.2f} م²   |   '
                    f'عدد السجلات: {len(records)}')
    elements.append(Paragraph(summary_text, subtitle_style))

    doc.build(elements)
    buf.seek(0)
    return buf


# ══════════════════════════════════════════════════════════════════
# 🛰️ SATELLITE REPORT
# ══════════════════════════════════════════════════════════════════
def _resolve_media_path(url_or_path):
    """Convert a media URL or path to an absolute local filesystem path."""
    from django.conf import settings
    if not url_or_path:
        return None
    # Already an absolute local path
    if os.path.isabs(url_or_path):
        return url_or_path
    # Strip MEDIA_URL prefix if present
    path = url_or_path
    mu = str(settings.MEDIA_URL)
    if path.startswith(mu):
        path = path[len(mu):]
    # Strip leading slash
    path = path.lstrip('/')
    # Normalize separators
    path = path.replace('/', os.sep)
    abs_path = os.path.join(settings.MEDIA_ROOT, path)
    return abs_path

def export_satellite_report(violation, years):
    """تقرير PDF مقارن باستخدام صور الأقمار الصناعية"""
    from django.conf import settings
    from .services import satellite as sat_svc

    buf = io.BytesIO()
    font_name = 'Helvetica'
    win_fonts = [
        'C:\\Windows\\Fonts\\arial.ttf',
        'C:\\Windows\\Fonts\\tahoma.ttf',
        'C:\\Windows\\Fonts\\times.ttf',
        'C:\\Windows\\Fonts\\Candarab.ttf',
    ]
    linux_fonts = [
        '/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf',
    ]
    for fp in win_fonts + linux_fonts:
        if os.path.exists(fp):
            try:
                pdfmetrics.registerFont(TTFont('ReportFont', fp))
                font_name = 'ReportFont'
                break
            except Exception:
                pass

    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        rightMargin=1.5*cm, leftMargin=1.5*cm,
        topMargin=2*cm, bottomMargin=2*cm,
        title=f'التقرير الفضائي — {violation.code}',
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('T', fontName=font_name, fontSize=14,
        textColor=colors.HexColor('#0D3B6E'), alignment=TA_CENTER, spaceAfter=6)
    sub_style = ParagraphStyle('S', fontName=font_name, fontSize=9,
        textColor=colors.grey, alignment=TA_CENTER, spaceAfter=4)
    h3_style = ParagraphStyle('H3', fontName=font_name, fontSize=10,
        textColor=colors.HexColor('#0D3B6E'), alignment=TA_LEFT, spaceAfter=6)
    normal = ParagraphStyle('N', fontName=font_name, fontSize=9, alignment=TA_LEFT)

    elements = []
    now = datetime.now().strftime('%Y-%m-%d %H:%M')

    # ── Header ──
    elements.append(Paragraph(ar('تقرير الرصد الفضائي — تواجدات طرح النهر'), title_style))
    elements.append(Paragraph(ar(f'وزارة الموارد المائية والري | {now}'), sub_style))
    elements.append(HRFlowable(width='100%', thickness=2,
        color=colors.HexColor('#0D3B6E'), spaceAfter=8))

    # ── Violation info ──
    gov = violation.governorate.name_ar if violation.governorate else '—'
    info_lines = [
        f'الرمز: {violation.code}',
        f'المحافظة: {gov} | المركز: {violation.district_name}',
        f'القرية: {violation.village} | المستغل: {violation.occupant}',
        f'الإحداثيات: {violation.latitude:.5f}, {violation.longitude:.5f} | '
        f'المساحة: {violation.area_total:,.1f} م²',
    ]
    for line in info_lines:
        elements.append(Paragraph(ar(line), normal))
    elements.append(Spacer(1, 0.3*cm))

    # ── Fetch satellite data ──
    lat, lng = violation.latitude, violation.longitude
    result = sat_svc.compute_change_detection(lat, lng, years)

    if 'error' in result:
        elements.append(Paragraph(ar(f'خطأ: {result["error"]}'), normal))
        doc.build(elements)
        buf.seek(0)
        return buf

    # ── Images Section ──
    elements.append(Paragraph(ar('الصور الفضائية'), h3_style))
    for img in result.get('images', []):
        img_url = img.get('image_url', '')
        abs_path = _resolve_media_path(img_url)
        if abs_path and os.path.exists(abs_path):
            img_w = 14*cm
            elements.append(Paragraph(ar(f'سنة {img["year"]}'), sub_style))
            elements.append(Image(abs_path, width=img_w, height=img_w*0.85))
            elements.append(Spacer(1, 0.2*cm))

    # ── Change Detection Section ──
    if result.get('changes'):
        elements.append(Paragraph(ar('كشف التغيير'), h3_style))
        for ch in result['changes']:
            elements.append(Paragraph(ar(f'{ch["year_from"]} ← {ch["year_to"]}'), sub_style))
            text = (
                f'متوسط الفرق: {ch["mean_diff"]} | '
                f'نسبة التغيير: {ch["pct_changed"]}% | '
                f'مساحة تقديرية: {ch["est_area_m2"]:,.1f} م²'
            )
            elements.append(Paragraph(ar(text), normal))
            hp_url = ch.get('heatmap_url', '')
            hp_path = _resolve_media_path(hp_url)
            if hp_path and os.path.exists(hp_path):
                elements.append(Image(hp_path, width=12*cm, height=12*cm*0.85))
            elements.append(Spacer(1, 0.2*cm))

    # ── Statistics Table ──
    if result.get('stats'):
        elements.append(Paragraph(ar('إحصائيات'), h3_style))
        s = result['stats']
        stats = [
            [ar('الفترة'), f'{s["first_year"]} ← {s["last_year"]}'],
            [ar('متوسط التغيير'), f'{s["avg_change_pct"]}%'],
            [ar('المساحة التقديرية الكلية'), f'{s["total_est_area_m2"]:,.1f} م²'],
        ]
        tbl = Table([[ar('البيان'), ar('القيمة')]] + stats,
                     colWidths=[8*cm, 8*cm])
        tbl.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0D3B6E')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,-1), font_name),
            ('FONTSIZE', (0,0), (-1,-1), 9),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('GRID', (0,0), (-1,-1), 0.4, colors.HexColor('#CCCCCC')),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#EEF4FB')]),
        ]))
        elements.append(tbl)

    # ── Footer ──
    elements.append(Spacer(1, 0.5*cm))
    elements.append(HRFlowable(width='100%', thickness=1,
        color=colors.HexColor('#CCCCCC'), spaceAfter=6))
    elements.append(Paragraph(ar(
        f'تم الإنشاء: {now} | المصدر: Sentinel-2 عبر Sentinel Hub'),
        sub_style))

    doc.build(elements)
    buf.seek(0)
    return buf
