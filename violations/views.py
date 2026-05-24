import io
import json
import os
import tempfile
import zipfile
from django.db import models
from django.conf import settings
from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.db.models import Sum, Count, Q
from django.utils import timezone
from .models import Violation, Governorate, District, UserProfile, ViolationImage, ViolationNote, ViolationUsage

_GEO_ALL = None
_GOVS    = None

def _load_geo():
    global _GEO_ALL, _GOVS
    if _GEO_ALL is None:
        with open(settings.GEO_JSON_PATH, 'r', encoding='utf-8') as f:
            _GEO_ALL = json.load(f)
    if _GOVS is None:
        with open(settings.GOVS_JSON_PATH, 'r', encoding='utf-8') as f:
            _GOVS = json.load(f)

def get_role(user):
    if not user.is_authenticated: return None
    if user.is_superuser: return 'manager'
    try: return user.profile.role
    except: return 'viewer'

# ── AUTH ──────────────────────────────────────────────────────────
def login_view(request):
    error = None
    if request.method == 'POST':
        username = request.POST.get('username','')
        u = authenticate(request, username=username,
                         password=request.POST.get('password',''))
        if u:
            login(request, u)
            ip = (request.META.get('HTTP_X_FORWARDED_FOR','').split(',')[0].strip()
                  or request.META.get('REMOTE_ADDR',''))
            from .models import AuditLog
            AuditLog.objects.create(user=u, action='login',
                target='تسجيل دخول ناجح', ip_address=ip or None)
            return redirect('map')
        error = 'اسم المستخدم أو كلمة المرور غير صحيحة'
        ip = (request.META.get('HTTP_X_FORWARDED_FOR','').split(',')[0].strip()
              or request.META.get('REMOTE_ADDR',''))
        from .models import AuditLog
        AuditLog.objects.create(user=None, action='login_fail',
            target=f'محاولة فاشلة: {username}', ip_address=ip or None)
    return render(request, 'violations/login.html', {'error': error})

def logout_view(request):
    if request.user.is_authenticated:
        log_action(request, 'logout', 'تسجيل خروج')
    logout(request)
    return redirect('login')

# ── MAIN MAP ──────────────────────────────────────────────────────
@login_required(login_url='login')
def map_view(request):
    role = get_role(request.user)
    govs_with_data = Governorate.objects.filter(has_data=True).order_by('name_ar')
    overall = Violation.objects.filter(status='approved').aggregate(
        total_count=Count('id'), total_area=Sum('area_total'),
        geo_count=Count('id', filter=Q(geo_exact=True)),
    )
    pending_count = Violation.objects.filter(status='pending').count() if role in ('supervisor','manager') else 0

    # كل محافظات مصر من ملف الجيو (بغض النظر عن وجود بيانات)
    _load_geo()
    all_egypt_govs = sorted(_GOVS, key=lambda g: g['name_ar'])

    context = {
        'govs_with_data': govs_with_data, 'overall': overall,
        'role': role, 'pending_count': pending_count,
        'all_govs': all_egypt_govs,
        'user': request.user,
    }
    return render(request, 'violations/map.html', context)

# ── GEO API ───────────────────────────────────────────────────────
@login_required(login_url='login')
def geo_gov_api(request, gov_pcode):
    _load_geo()
    polys = [v for v in _GEO_ALL if v['gov_pcode'] == gov_pcode]
    counts = dict(Violation.objects.filter(governorate__pcode=gov_pcode, status='approved')
                  .values_list('village_pcode').annotate(c=Count('id')))
    areas  = dict(Violation.objects.filter(governorate__pcode=gov_pcode, status='approved')
                  .values_list('village_pcode').annotate(s=Sum('area_total')))
    for p in polys:
        p['violation_count'] = counts.get(p['pcode'], 0)
        p['total_area']      = round(areas.get(p['pcode'], 0) or 0, 1)
    return JsonResponse({'polygons': polys, 'gov_pcode': gov_pcode})

# ── VIOLATIONS API ────────────────────────────────────────────────
@login_required(login_url='login')
def violations_api(request):
    role = get_role(request.user)
    qs = Violation.objects.select_related('governorate').all()
    if role not in ('supervisor', 'manager'):
        qs = qs.filter(status='approved')

    gov=request.GET.get('gov',''); district=request.GET.get('district','')
    village=request.GET.get('village',''); desc=request.GET.get('description','')
    min_area=request.GET.get('min_area',''); search=request.GET.get('search','')
    pcode=request.GET.get('pcode',''); status=request.GET.get('status','')

    if gov:      qs = qs.filter(governorate__pcode=gov)
    if district: qs = qs.filter(district_name=district)
    if village:  qs = qs.filter(village=village)
    if desc:     qs = qs.filter(description__icontains=desc)
    if pcode:    qs = qs.filter(village_pcode=pcode)
    if status and role in ('supervisor','manager'): qs = qs.filter(status=status)
    if min_area:
        try: qs = qs.filter(area_total__gte=float(min_area))
        except: pass
    if search:
        qs = qs.filter(Q(occupant__icontains=search)|Q(code__icontains=search)|
                       Q(village__icontains=search)|Q(basin__icontains=search))

    summary = qs.aggregate(count=Count('id'), total_area=Sum('area_total'))

    # فلتر no_geometry — يُعيد فقط السجلات التي لها lat/lon ولا geometry (للعرض كـ circles)
    no_geo = request.GET.get('no_geometry', '')
    if no_geo:
        qs = qs.filter(geometry__isnull=True).exclude(latitude=0, longitude=0)

    records = list(qs.values('id','code','governorate__name_ar','governorate__pcode',
        'district_name','village','village_pcode','occupant','basin','description',
        'area_outside','area_public','area_nile_bank','area_total',
        'latitude','longitude','geo_exact','status','field_notes')[:2000])
    return JsonResponse({'records': records, 'summary': summary, 'role': role})

# ── FILTER OPTIONS ────────────────────────────────────────────────
@login_required(login_url='login')
def filter_options_api(request):
    gov      = request.GET.get('gov', '')
    district = request.GET.get('district', '')

    # Districts: from geo JSON (all villages) + DB (existing violations)
    _load_geo()
    geo_districts = []
    geo_villages  = []

    if gov and _GEO_ALL:
        # Get districts from geo data
        dist_set = set()
        for v in _GEO_ALL:
            if v['gov_pcode'] == gov:
                dist_set.add(v['district_ar'])
        geo_districts = sorted(dist_set)

        # Get villages filtered by district
        if district:
            for v in _GEO_ALL:
                if v['gov_pcode'] == gov and v['district_ar'] == district:
                    geo_villages.append(v['name_ar'])
            geo_villages = sorted(geo_villages)

    # Also get existing violation descriptions
    qs = Violation.objects.all()
    if gov:      qs = qs.filter(governorate__pcode=gov)
    if district: qs = qs.filter(district_name=district)
    descs = list(qs.values_list('description', flat=True).distinct().order_by('description'))

    # Merge DB districts with geo districts
    db_districts = list(qs.values_list('district_name', flat=True).distinct())
    all_districts = sorted(set(geo_districts + db_districts))

    # Merge DB villages with geo villages
    db_villages = []
    if district:
        db_villages = list(qs.filter(district_name=district)
                           .values_list('village', flat=True).distinct())
    all_villages = sorted(set(geo_villages + db_villages))

    return JsonResponse({
        'districts':    all_districts,
        'villages':     all_villages,
        'descriptions': descs,
    })

# ── GOVS SUMMARY ──────────────────────────────────────────────────
@login_required(login_url='login')
def govs_summary_api(request):
    stats = (Violation.objects.filter(status='approved')
             .values('governorate__pcode','governorate__name_ar','governorate__name_en')
             .annotate(count=Count('id'),total_area=Sum('area_total')).order_by('-count'))
    return JsonResponse({'governorates': list(stats)})

# ══════════════════════════════════════════════════════════════════
# SHAPEFILE UPLOAD & IMPORT
# ══════════════════════════════════════════════════════════════════

# خريطة حقول السمات الشائعة → حقول النموذج
# يمكن توسيعها لدعم أسماء حقول مختلفة في ملفات الشيب فايل
_FIELD_MAP = {
    # occupant
    'occupant':    'occupant',  'مستغل':      'occupant',  'owner':       'occupant',
    'name':        'occupant',  'الاسم':      'occupant',  'اسم_المستغل': 'occupant',
    'اسم المستغل': 'occupant',
    # basin
    'basin':       'basin',     'حوض':        'basin',     'basin_name':  'basin',
    'اسم الحوض':  'basin',     'اسم_الحوض':  'basin',
    # description
    'description': 'description', 'وصف':      'description', 'desc':      'description',
    'وصف_الاستغلال': 'description', 'وصف الاستغلال': 'description',
    'type':        'description', 'نوع':       'description',
    # district
    'district':    'district_name', 'مركز':   'district_name', 'district_name': 'district_name',
    'اسم المركز': 'district_name',  'اسم_المركز': 'district_name',
    # village
    'village':     'village',   'قرية':       'village',   'village_name': 'village',
    'القرية':      'village',   'اسم_القرية': 'village',   'اسم القرية':  'village',
    'city':        'village',
    # areas
    'area_outside':   'area_outside',   'خارج_الحياض':    'area_outside',
    'area_public':    'area_public',    'منفعة_عامة':     'area_public',
    'area_nile_bank': 'area_nile_bank', 'جسر_النهر':      'area_nile_bank',
    'area_total':     'area_total',     'area':            'area_total',
    'المساحة':        'area_total',     'مساحة':           'area_total',
    'total_area':     'area_total',
    # code
    'code':        'code',      'رمز':        'code',      'fid':         'code',
    'id':          'code',      'ref':         'code',
    # gov
    'gov':         'gov_name',  'governorate': 'gov_name',  'محافظة':     'gov_name',
    'gov_name':    'gov_name',  'اسم_المحافظة': 'gov_name',
}

GOV_PCODE_MAP = {
    'القاهرة':'EG01','الاسكندرية':'EG02','الإسكندرية':'EG02','بورسعيد':'EG03',
    'السويس':'EG04','دمياط':'EG11','الدقهلية':'EG12','الشرقية':'EG13',
    'القليوبية':'EG14','كفر الشيخ':'EG15','الغربية':'EG16','المنوفية':'EG17',
    'البحيرة':'EG18','الإسماعيلية':'EG19','الجيزة':'EG21','بنى سويف':'EG22',
    'بني سويف':'EG22','الفيوم':'EG23','المنيا':'EG24','أسيوط':'EG25',
    'اسيوط':'EG25','سوهاج':'EG26','قنا':'EG27','أسوان':'EG28','اسوان':'EG28',
    'الأقصر':'EG29','مدينة الأقصر':'EG29','البحر الأحمر':'EG31',
    'الوادى الجديد':'EG32','الوادي الجديد':'EG32','مطروح':'EG33',
    'شمال سيناء':'EG34','جنوب سيناء':'EG35',
}


def _safe_float(val, default=0.0):
    try:
        return float(val) if val is not None else default
    except (TypeError, ValueError):
        return default


def _safe_str(val, default=''):
    if val is None:
        return default
    return str(val).strip()


def _map_row_fields(row_dict):
    """يحوّل حقول السمات إلى الحقول المناسبة في النموذج."""
    result = {}
    for raw_key, raw_val in row_dict.items():
        normalized = str(raw_key).strip().lower().replace('-', '_')
        mapped = _FIELD_MAP.get(normalized) or _FIELD_MAP.get(str(raw_key).strip())
        if mapped:
            result[mapped] = raw_val
    return result


@login_required(login_url='login')
def shapefile_preview_api(request):
    """
    يستقبل ملف shapefile.zip ويعيد:
    - قائمة حقول السمات
    - عدد القطع
    - نموذج للسجلات الأولى
    - اقتراح ربط الحقول
    """
    role = get_role(request.user)
    if role not in ('data_entry', 'supervisor', 'manager'):
        return JsonResponse({'error': 'غير مصرح'}, status=403)
    if request.method != 'POST':
        return JsonResponse({'error': 'POST only'}, status=405)

    shp_file = request.FILES.get('shapefile')
    if not shp_file:
        return JsonResponse({'error': 'لم يتم رفع ملف'}, status=400)

    try:
        import geopandas as gpd
        from shapely.geometry import mapping

        with tempfile.TemporaryDirectory() as tmpdir:
            zip_path = os.path.join(tmpdir, 'upload.zip')
            with open(zip_path, 'wb') as f:
                for chunk in shp_file.chunks():
                    f.write(chunk)

            with zipfile.ZipFile(zip_path, 'r') as z:
                z.extractall(tmpdir)

            # البحث عن ملف .shp داخل المجلد
            shp_path = None
            for root, dirs, files in os.walk(tmpdir):
                for fname in files:
                    if fname.lower().endswith('.shp'):
                        shp_path = os.path.join(root, fname)
                        break
                if shp_path:
                    break

            if not shp_path:
                return JsonResponse({'error': 'لم يتم العثور على ملف .shp داخل الأرشيف'}, status=400)

            gdf = gpd.read_file(shp_path)
            if gdf.crs and gdf.crs.to_epsg() != 4326:
                gdf = gdf.to_crs(epsg=4326)

            fields = [c for c in gdf.columns if c != 'geometry']
            total  = len(gdf)

            # اقتراح ربط الحقول تلقائياً
            suggested = {}
            for f in fields:
                normalized = f.strip().lower().replace('-', '_')
                mapped = _FIELD_MAP.get(normalized) or _FIELD_MAP.get(f.strip())
                if mapped:
                    suggested[f] = mapped

            # نموذج السجلات الأولى (5 فقط)
            samples = []
            for _, row in gdf.head(5).iterrows():
                s = {}
                for f in fields:
                    val = row.get(f)
                    s[f] = _safe_str(val) if val is not None else ''
                samples.append(s)

            return JsonResponse({
                'fields':    fields,
                'total':     total,
                'suggested': suggested,
                'samples':   samples,
            })

    except Exception as e:
        import traceback
        return JsonResponse({'error': str(e), 'trace': traceback.format_exc()}, status=500)


@login_required(login_url='login')
def shapefile_import_api(request):
    """
    يستقبل ملف shapefile.zip + mapping الحقول + pcode المحافظة
    ويستورد جميع قطع أراضي طرح النهر مع الهندسة الجغرافية الكاملة.
    """
    role = get_role(request.user)
    if role not in ('data_entry', 'supervisor', 'manager'):
        return JsonResponse({'error': 'غير مصرح'}, status=403)
    if request.method != 'POST':
        return JsonResponse({'error': 'POST only'}, status=405)

    shp_file  = request.FILES.get('shapefile')
    gov_pcode = request.POST.get('gov_pcode', '').strip()
    field_map_json = request.POST.get('field_map', '{}')
    clear_gov = request.POST.get('clear_gov', 'false').lower() == 'true'

    if not shp_file:
        return JsonResponse({'error': 'لم يتم رفع ملف'}, status=400)
    if not gov_pcode:
        return JsonResponse({'error': 'كود المحافظة مطلوب'}, status=400)

    try:
        field_map = json.loads(field_map_json)
    except Exception:
        field_map = {}

    try:
        import geopandas as gpd
        from shapely.geometry import mapping, MultiPolygon, Polygon

        with tempfile.TemporaryDirectory() as tmpdir:
            zip_path = os.path.join(tmpdir, 'upload.zip')
            with open(zip_path, 'wb') as f:
                for chunk in shp_file.chunks():
                    f.write(chunk)

            with zipfile.ZipFile(zip_path, 'r') as z:
                z.extractall(tmpdir)

            shp_path = None
            for root, dirs, files in os.walk(tmpdir):
                for fname in files:
                    if fname.lower().endswith('.shp'):
                        shp_path = os.path.join(root, fname)
                        break
                if shp_path:
                    break

            if not shp_path:
                return JsonResponse({'error': 'لم يتم العثور على ملف .shp'}, status=400)

            gdf = gpd.read_file(shp_path)
            if gdf.crs and gdf.crs.to_epsg() != 4326:
                gdf = gdf.to_crs(epsg=4326)

        # ── المحافظة ──────────────────────────────────────────────
        _load_geo()
        gov = Governorate.objects.filter(pcode=gov_pcode).first()
        if not gov:
            gov_info = next((g for g in (_GOVS or []) if g.get('pcode') == gov_pcode), None)
            gov_name_ar = gov_info['name_ar'] if gov_info else gov_pcode
            gov_name_en = gov_info.get('name_en', gov_pcode) if gov_info else gov_pcode
            gov = Governorate.objects.create(
                pcode=gov_pcode, name_ar=gov_name_ar,
                name_en=gov_name_en, has_data=True,
            )
        if not gov.has_data:
            gov.has_data = True
            gov.save(update_fields=['has_data'])

        if clear_gov:
            Violation.objects.filter(governorate=gov).delete()

        # ── الاستيراد ─────────────────────────────────────────────
        import_batch = f'shp-{gov_pcode}-{timezone.now().strftime("%Y%m%d%H%M%S")}'
        v_status     = 'approved' if role in ('supervisor', 'manager') else 'pending'
        created = 0
        errors  = []
        start_num = Violation.objects.filter(governorate=gov).count()

        for idx, row in gdf.iterrows():
            try:
                # --- تطبيق field_map المستخدم ---
                attrs = {}
                for shp_col, model_field in field_map.items():
                    val = row.get(shp_col)
                    if val is not None:
                        attrs[model_field] = val

                # --- احتياطي: تطبيق _FIELD_MAP التلقائي ---
                auto = _map_row_fields({c: row.get(c) for c in gdf.columns if c != 'geometry'})
                for k, v in auto.items():
                    if k not in attrs:
                        attrs[k] = v

                # --- استنتاج المحافظة من السمات إن وُجدت ---
                gov_name_attr = _safe_str(attrs.pop('gov_name', ''))
                if gov_name_attr and gov_name_attr in GOV_PCODE_MAP:
                    inferred_pcode = GOV_PCODE_MAP[gov_name_attr]
                    if inferred_pcode != gov_pcode:
                        # قطعة تنتمي لمحافظة مختلفة — أنشئ/احضر تلك المحافظة
                        alt_gov = Governorate.objects.filter(pcode=inferred_pcode).first()
                        if not alt_gov:
                            alt_info = next((g for g in (_GOVS or []) if g.get('pcode') == inferred_pcode), None)
                            alt_gov  = Governorate.objects.create(
                                pcode=inferred_pcode,
                                name_ar=alt_info['name_ar'] if alt_info else gov_name_attr,
                                name_en=alt_info.get('name_en', inferred_pcode) if alt_info else inferred_pcode,
                                has_data=True,
                            )
                        current_gov = alt_gov
                    else:
                        current_gov = gov
                else:
                    current_gov = gov

                # --- هندسة القطعة ---
                geom = row.geometry
                geom_json = None
                centroid_lat = centroid_lon = 0.0
                if geom and not geom.is_empty:
                    geom_json = mapping(geom)
                    centroid   = geom.centroid
                    centroid_lat = centroid.y
                    centroid_lon = centroid.x

                # --- الحقول الأساسية ---
                num  = start_num + created + 1
                code = _safe_str(attrs.get('code')) or f"{current_gov.pcode}-{num:04d}"
                # تجنب الكود المكرر
                if Violation.objects.filter(code=code).exists():
                    code = f"{current_gov.pcode}-{num:04d}-{idx}"

                district_name = _safe_str(attrs.get('district_name')) or current_gov.name_ar
                village       = _safe_str(attrs.get('village'))       or district_name

                # --- حساب المساحة الإجمالية ---
                area_outside   = _safe_float(attrs.get('area_outside'))
                area_public    = _safe_float(attrs.get('area_public'))
                area_nile_bank = _safe_float(attrs.get('area_nile_bank'))
                area_total     = _safe_float(attrs.get('area_total'))
                if area_total == 0:
                    area_total = area_outside + area_public + area_nile_bank

                Violation.objects.create(
                    governorate    = current_gov,
                    district_name  = district_name,
                    village        = village,
                    village_pcode  = '',
                    code           = code,
                    occupant       = _safe_str(attrs.get('occupant'))    or '—',
                    basin          = _safe_str(attrs.get('basin'))        or '—',
                    description    = _safe_str(attrs.get('description'))  or '—',
                    area_outside   = area_outside,
                    area_public    = area_public,
                    area_nile_bank = area_nile_bank,
                    area_total     = area_total,
                    latitude       = centroid_lat,
                    longitude      = centroid_lon,
                    geo_exact      = geom is not None,
                    geometry       = geom_json,
                    field_notes    = '',
                    status         = v_status,
                    submitted_by   = request.user,
                    import_batch   = import_batch,
                )
                created += 1

            except Exception as e:
                errors.append(f'سجل {idx}: {str(e)}')

        log_action(request, 'import', f'Shapefile — {gov.name_ar} — {created} قطعة')
        return JsonResponse({
            'success': True,
            'created': created,
            'errors':  errors[:20],
            'message': f'تم استيراد {created} قطعة بنجاح من محافظة {gov.name_ar}',
        })

    except Exception as e:
        import traceback
        return JsonResponse({'error': str(e), 'trace': traceback.format_exc()}, status=500)


@login_required(login_url='login')
def violations_geojson_api(request):
    """
    يعيد GeoJSON لقطع الأراضي التي لها هندسة (polygon) — للعرض المباشر على الخريطة.
    """
    role = get_role(request.user)
    qs = Violation.objects.select_related('governorate').exclude(geometry__isnull=True)
    if role not in ('supervisor', 'manager'):
        qs = qs.filter(status='approved')

    gov      = request.GET.get('gov', '')
    district = request.GET.get('district', '')
    village  = request.GET.get('village', '')
    desc     = request.GET.get('description', '')
    min_area = request.GET.get('min_area', '')
    search   = request.GET.get('search', '')
    status_f = request.GET.get('status', '')
    pcode    = request.GET.get('pcode', '')

    if gov:      qs = qs.filter(governorate__pcode=gov)
    if district: qs = qs.filter(district_name=district)
    if village:  qs = qs.filter(village=village)
    if desc:     qs = qs.filter(description__icontains=desc)
    if pcode:    qs = qs.filter(village_pcode=pcode)
    if status_f and role in ('supervisor', 'manager'):
        qs = qs.filter(status=status_f)
    if min_area:
        try:
            qs = qs.filter(area_total__gte=float(min_area))
        except ValueError:
            pass
    if search:
        qs = qs.filter(
            Q(occupant__icontains=search) | Q(code__icontains=search) |
            Q(village__icontains=search)  | Q(basin__icontains=search)
        )

    features = []
    for v in qs[:2000]:
        gov_name  = v.governorate.name_ar if v.governorate else ''
        gov_pcode = v.governorate.pcode   if v.governorate else ''
        features.append({
            'type': 'Feature',
            'geometry': v.geometry,
            'properties': {
                'id':           v.id,
                'code':         v.code,
                'gov_name':     gov_name,
                'gov_pcode':    gov_pcode,
                'district_name':v.district_name,
                'village':      v.village,
                'occupant':     v.occupant,
                'basin':        v.basin,
                'description':  v.description,
                'area_outside': v.area_outside,
                'area_public':  v.area_public,
                'area_nile_bank': v.area_nile_bank,
                'area_total':   v.area_total,
                'status':       v.status,
                'geo_exact':    v.geo_exact,
            },
        })

    return JsonResponse({'type': 'FeatureCollection', 'features': features})

# ── EDIT VIOLATION ────────────────────────────────────────────────
@login_required(login_url='login')
def edit_violation_api(request, pk):
    role = get_role(request.user)
    if role not in ('data_entry','supervisor','manager'):
        return JsonResponse({'error':'غير مصرح'},status=403)
    v = get_object_or_404(Violation, pk=pk)
    if role=='data_entry' and (v.submitted_by!=request.user or v.status!='pending'):
        return JsonResponse({'error':'لا يمكنك تعديل هذا السجل'},status=403)
    try: data = json.loads(request.body)
    except: return JsonResponse({'error':'بيانات غير صالحة'},status=400)

    for f in ['district_name','village','occupant','basin','description','field_notes']:
        if f in data: setattr(v, f, str(data[f]).strip())
    for f in ['area_outside','area_public','area_nile_bank','area_total','latitude','longitude']:
        if f in data: setattr(v, f, float(data[f] or 0))
    if role=='data_entry': v.status='pending'
    v.save()
    return JsonResponse({'success':True,'message':'تم التعديل بنجاح'})

# ── APPROVE / REJECT ──────────────────────────────────────────────
@login_required(login_url='login')
def approve_violation_api(request, pk):
    role = get_role(request.user)
    if role not in ('supervisor','manager'):
        return JsonResponse({'error':'غير مصرح'},status=403)
    v = get_object_or_404(Violation, pk=pk)
    try: data = json.loads(request.body)
    except: data = {}
    v.reviewed_by  = request.user
    v.reviewed_at  = timezone.now()
    v.review_notes = data.get('notes','')
    v.status       = 'approved' if data.get('action','approve')=='approve' else 'rejected'
    v.save()
    msg = 'تم اعتماد السجل' if v.status=='approved' else 'تم رفض السجل'
    return JsonResponse({'success':True,'message':msg,'status':v.status})

# ── NOTES ─────────────────────────────────────────────────────────
@login_required(login_url='login')
def notes_api(request, pk):
    v = get_object_or_404(Violation, pk=pk)
    if request.method=='GET':
        notes = list(v.notes.select_related('user').values(
            'id','text','user__username','user__first_name','created_at'))
        return JsonResponse({'notes':notes})
    if request.method=='POST':
        try: data = json.loads(request.body)
        except: return JsonResponse({'error':'بيانات غير صالحة'},status=400)
        text = data.get('text','').strip()
        if not text: return JsonResponse({'error':'الملاحظة فارغة'},status=400)
        note = ViolationNote.objects.create(violation=v,user=request.user,text=text)
        return JsonResponse({'success':True,'id':note.id,'text':note.text,
            'user':request.user.get_full_name() or request.user.username,
            'created_at':note.created_at.strftime('%Y-%m-%d %H:%M')})
    return JsonResponse({'error':'Method not allowed'},status=405)

# ── IMAGE UPLOAD ──────────────────────────────────────────────────
@login_required(login_url='login')
def upload_image_api(request, pk):
    role = get_role(request.user)
    if role not in ('data_entry','supervisor','manager'):
        return JsonResponse({'error':'غير مصرح'},status=403)
    if request.method!='POST':
        return JsonResponse({'error':'POST only'},status=405)
    v = get_object_or_404(Violation, pk=pk)
    images = request.FILES.getlist('images')
    if not images: return JsonResponse({'error':'لم يتم رفع أي صورة'},status=400)
    saved = []
    for img in images:
        vi = ViolationImage.objects.create(
            violation=v, image=img,
            caption=request.POST.get('caption',''),
            uploaded_by=request.user)
        saved.append({'id':vi.id,'url':vi.image.url})
    return JsonResponse({'success':True,'images':saved})


# ── SERVE SURVEY MAP (مع headers صحيحة لـ PDF.js) ────────────────
@login_required(login_url='login')
def serve_survey_map(request, pk):
    """يخدم ملف PDF مع headers تسمح لـ PDF.js بقراءته"""
    from django.http import FileResponse, Http404
    v = get_object_or_404(Violation, pk=pk)
    if not v.survey_map:
        raise Http404
    try:
        response = FileResponse(
            open(v.survey_map.path, 'rb'),
            content_type='application/pdf'
        )
        response['Content-Disposition'] = f'inline; filename="{v.code}.pdf"'
        response['Access-Control-Allow-Origin'] = '*'
        response['X-Frame-Options'] = 'SAMEORIGIN'
        return response
    except FileNotFoundError:
        raise Http404

# ── SURVEY MAP UPLOAD ──────────────────────────────────────────────
@login_required(login_url='login')
def upload_survey_map_api(request, pk):
    """رفع الخريطة المساحية الرسمية (PDF) وربطها بقطعة الأرض"""
    role = get_role(request.user)
    if role not in ('data_entry', 'supervisor', 'manager'):
        return JsonResponse({'error': 'غير مصرح'}, status=403)
    if request.method != 'POST':
        return JsonResponse({'error': 'POST only'}, status=405)

    v   = get_object_or_404(Violation, pk=pk)
    pdf = request.FILES.get('survey_map')

    if not pdf:
        return JsonResponse({'error': 'لم يتم اختيار ملف'}, status=400)
    if not pdf.name.lower().endswith('.pdf'):
        return JsonResponse({'error': 'يجب أن يكون الملف بصيغة PDF'}, status=400)
    if pdf.size > 20 * 1024 * 1024:   # 20 MB
        return JsonResponse({'error': 'حجم الملف يتجاوز 20 ميجابايت'}, status=400)

    # حذف الملف القديم إن وُجد
    if v.survey_map:
        try:
            import os
            if os.path.isfile(v.survey_map.path):
                os.remove(v.survey_map.path)
        except Exception:
            pass

    v.survey_map = pdf
    v.save(update_fields=['survey_map'])

    log_action(request, 'edit',
               f'رفع خريطة مساحية للقطعة: {v.code}')
    return JsonResponse({
        'success': True,
        'url':     f'/api/violations/{v.pk}/survey-map/view/',
        'name':    pdf.name,
        'message': 'تم رفع الخريطة المساحية بنجاح',
    })

# ── VIOLATION DETAIL ──────────────────────────────────────────────
@login_required(login_url='login')
def violation_detail_api(request, pk):
    import traceback as tb
    role = get_role(request.user)

    try:
        v = get_object_or_404(Violation, pk=pk)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=404)

    # safe getter
    def sg(obj, attr, default=''):
        try:    return getattr(obj, attr) or default
        except: return default

    def sf(obj, attr, default=0.0):
        try:    return float(getattr(obj, attr) or 0)
        except: return default

    try:
        status_val = sg(v, 'status', 'approved')
        if status_val == 'pending' and role not in ('supervisor', 'manager'):
            try:
                if v.submitted_by != request.user:
                    return JsonResponse({'error': 'غير مصرح'}, status=403)
            except Exception:
                pass

        try:
            images = [
                {'id': i.id, 'url': i.image.url, 'caption': sg(i, 'caption')}
                for i in v.images.all()
            ]
        except Exception:
            images = []

        try:
            sub_by = v.submitted_by.get_full_name() if v.submitted_by else ''
        except Exception:
            sub_by = ''

        try:
            sub_at = v.submitted_at.strftime('%Y-%m-%d %H:%M') if v.submitted_at else ''
        except Exception:
            sub_at = ''

        try:
            gov_pcode = v.governorate.pcode if v.governorate else ''
            gov_name  = v.governorate.name_ar if v.governorate else ''
        except Exception:
            gov_pcode = gov_name = ''

        data = {
            'id':           v.id,
            'code':         sg(v, 'code'),
            'status':       status_val,
            'gov_pcode':    gov_pcode,
            'gov_name':     gov_name,
            'district_name':sg(v, 'district_name'),
            'village':      sg(v, 'village'),
            'village_pcode':sg(v, 'village_pcode'),
            'occupant':     sg(v, 'occupant'),
            'basin':        sg(v, 'basin'),
            'description':  sg(v, 'description'),
            'area_outside': sf(v, 'area_outside'),
            'area_public':  sf(v, 'area_public'),
            'area_nile_bank': sf(v, 'area_nile_bank'),
            'area_total':   sf(v, 'area_total'),
            'latitude':     sf(v, 'latitude'),
            'longitude':    sf(v, 'longitude'),
            'geo_exact':    bool(sg(v, 'geo_exact', False)),
            'geometry':     v.geometry,
            'survey_map_url':  f'/api/violations/{v.pk}/survey-map/view/' if v.survey_map else None,
            'survey_map_name': v.survey_map.name.split('/')[-1] if v.survey_map else None,
            'field_notes':  sg(v, 'field_notes'),
            'submitted_by': sub_by,
            'submitted_at': sub_at,
            'review_notes': sg(v, 'review_notes'),
            'images':       images,
            'notes': [],
            'role':         role,
        }
        return JsonResponse(data)

    except Exception as e:
        return JsonResponse({
            'error': str(e),
            'trace': tb.format_exc()
        }, status=500)

# ── PENDING LIST ──────────────────────────────────────────────────
@login_required(login_url='login')
def pending_api(request):
    role = get_role(request.user)
    if role not in ('supervisor','manager'):
        return JsonResponse({'error':'غير مصرح'},status=403)
    qs = Violation.objects.filter(status='pending').select_related('governorate','submitted_by')
    records = list(qs.values('id','code','governorate__name_ar','district_name','village',
        'occupant','description','area_total','submitted_by__username',
        'submitted_by__first_name','submitted_at'))
    return JsonResponse({'records':records,'count':len(records)})


# ══════════════════════════════════════════════════════════════════
# AUDIT LOG HELPER
# ══════════════════════════════════════════════════════════════════
def log_action(request, action, target='', details=''):
    from .models import AuditLog
    ip = (request.META.get('HTTP_X_FORWARDED_FOR','').split(',')[0].strip()
          or request.META.get('REMOTE_ADDR',''))
    AuditLog.objects.create(
        user=request.user if request.user.is_authenticated else None,
        action=action, target=target, details=details, ip_address=ip or None,
    )


# ══════════════════════════════════════════════════════════════════
# ADMIN DASHBOARD
# ══════════════════════════════════════════════════════════════════
@login_required(login_url='login')
def admin_dashboard(request):
    role = get_role(request.user)
    if role != 'manager':
        return redirect('map')
    log_action(request, 'login', 'لوحة الإدارة')
    return render(request, 'violations/admin_dashboard.html', {'user': request.user})


@login_required(login_url='login')
def admin_stats_api(request):
    if get_role(request.user) != 'manager':
        return JsonResponse({'error': 'غير مصرح'}, status=403)
    from .models import AuditLog
    from django.utils import timezone
    today = timezone.now().date()
    stats = {
        'total':      Violation.objects.count(),
        'approved':   Violation.objects.filter(status='approved').count(),
        'pending':    Violation.objects.filter(status='pending').count(),
        'rejected':   Violation.objects.filter(status='rejected').count(),
        'total_area': Violation.objects.aggregate(s=Sum('area_total'))['s'] or 0,
        'geo_count':  Violation.objects.filter(geo_exact=True).count(),
        'user_count': User.objects.filter(is_active=True).count(),
        'gov_count':  Governorate.objects.filter(has_data=True).count(),
        'today_logs': AuditLog.objects.filter(timestamp__date=today).count(),
        'top_govs': list(
            Violation.objects.filter(status='approved')
            .values('governorate__name_ar')
            .annotate(count=Count('id'))
            .order_by('-count')[:5]
            .values('count', name=models.F('governorate__name_ar'))
        ),
    }
    return JsonResponse(stats)


@login_required(login_url='login')
def admin_users_api(request, user_id=None):
    if get_role(request.user) != 'manager':
        return JsonResponse({'error': 'غير مصرح'}, status=403)

    if request.method == 'GET':
        if user_id:
            u = get_object_or_404(User, pk=user_id)
            try:    role = u.profile.role
            except: role = 'viewer'
            try:    gov_pcode = u.profile.governorate.pcode if u.profile.governorate else ''
            except: gov_pcode = ''
            return JsonResponse({
                'id': u.id, 'username': u.username,
                'first_name': u.first_name, 'last_name': u.last_name,
                'email': u.email, 'role': role, 'gov_pcode': gov_pcode,
            })
        users = User.objects.select_related('profile__governorate').all().order_by('username')
        data = []
        for u in users:
            try:    role = u.profile.role
            except: role = 'viewer'
            try:    gov = u.profile.governorate.name_ar if u.profile.governorate else ''
            except: gov = ''
            data.append({
                'id': u.id, 'username': u.username,
                'full_name': u.get_full_name(), 'email': u.email,
                'role': role, 'governorate': gov,
                'is_active': u.is_active,
                'last_login': u.last_login.isoformat() if u.last_login else None,
            })
        return JsonResponse({'users': data})

    if request.method == 'POST':
        try: data = json.loads(request.body)
        except: return JsonResponse({'error': 'بيانات غير صالحة'}, status=400)

        if user_id:
            u = get_object_or_404(User, pk=user_id)
            u.first_name = data.get('first_name', u.first_name)
            u.last_name  = data.get('last_name',  u.last_name)
            u.email      = data.get('email',       u.email)
            if data.get('password'): u.set_password(data['password'])
            u.save()
        else:
            if not data.get('username'):
                return JsonResponse({'error': 'اسم المستخدم مطلوب'}, status=400)
            u, created = User.objects.get_or_create(username=data['username'])
            u.first_name = data.get('first_name', '')
            u.last_name  = data.get('last_name', '')
            u.email      = data.get('email', '')
            if data.get('password'): u.set_password(data['password'])
            elif created: u.set_password('Change@123')
            u.save()

        profile, _ = UserProfile.objects.get_or_create(user=u)
        profile.role = data.get('role', 'viewer')
        gov_pcode = data.get('gov_pcode', '')
        if gov_pcode:
            try: profile.governorate = Governorate.objects.get(pcode=gov_pcode)
            except: profile.governorate = None
        else:
            profile.governorate = None
        profile.save()
        log_action(request, 'edit', f'مستخدم: {u.username}')
        return JsonResponse({'success': True, 'message': f'تم حفظ المستخدم {u.username}'})


@login_required(login_url='login')
def admin_toggle_user_api(request, user_id):
    if get_role(request.user) != 'manager':
        return JsonResponse({'error': 'غير مصرح'}, status=403)
    u = get_object_or_404(User, pk=user_id)
    if u == request.user:
        return JsonResponse({'error': 'لا يمكنك إيقاف حسابك'}, status=400)
    u.is_active = not u.is_active
    u.save()
    msg = f'تم {"تفعيل" if u.is_active else "إيقاف"} المستخدم {u.username}'
    log_action(request, 'edit', msg)
    return JsonResponse({'success': True, 'message': msg})


@login_required(login_url='login')
def admin_logs_api(request):
    if get_role(request.user) != 'manager':
        return JsonResponse({'error': 'غير مصرح'}, status=403)
    from .models import AuditLog
    qs = AuditLog.objects.select_related('user').all()
    action = request.GET.get('action', '')
    user   = request.GET.get('user', '')
    date   = request.GET.get('date', '')
    limit  = int(request.GET.get('limit', 100))
    if action: qs = qs.filter(action=action)
    if user:   qs = qs.filter(user__username=user)
    if date:
        try:
            from datetime import date as dt
            d = dt.fromisoformat(date)
            qs = qs.filter(timestamp__date=d)
        except: pass
    logs = list(qs[:limit].values(
        'id','action','target','details','ip_address','timestamp',
        'user__username','user__first_name',
    ))
    users = list(AuditLog.objects.values_list('user__username', flat=True)
                 .distinct().exclude(user__username=None))
    return JsonResponse({
        'logs': [{
            'action': l['action'],
            'action_display': dict(AuditLog.ACTION_CHOICES).get(l['action'], l['action']),
            'target': l['target'],
            'details': l['details'],
            'ip': l['ip_address'],
            'user': l['user__first_name'] or l['user__username'] or 'غير معروف',
            'time': l['timestamp'].strftime('%Y-%m-%d %H:%M') if l['timestamp'] else '',
        } for l in logs],
        'users': users,
    })


@login_required(login_url='login')
def admin_logs_export(request):
    if get_role(request.user) != 'manager':
        return JsonResponse({'error': 'غير مصرح'}, status=403)
    from .models import AuditLog
    import openpyxl
    from django.http import HttpResponse
    import io
    qs = AuditLog.objects.select_related('user').all()[:5000]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'سجل الأنشطة'
    ws.sheet_view.rightToLeft = True
    headers = ['التوقيت','المستخدم','الحدث','الهدف','التفاصيل','عنوان IP']
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    action_labels = dict(AuditLog.ACTION_CHOICES)
    for row_i, log in enumerate(qs, 2):
        ws.cell(row=row_i, column=1, value=log.timestamp.strftime('%Y-%m-%d %H:%M') if log.timestamp else '')
        ws.cell(row=row_i, column=2, value=log.user.username if log.user else 'غير معروف')
        ws.cell(row=row_i, column=3, value=action_labels.get(log.action, log.action))
        ws.cell(row=row_i, column=4, value=log.target)
        ws.cell(row=row_i, column=5, value=log.details)
        ws.cell(row=row_i, column=6, value=log.ip_address or '')
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    response = HttpResponse(buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="audit_log.xlsx"'
    log_action(request, 'export', 'سجل الأنشطة')
    return response


@login_required(login_url='login')
def admin_govs_api(request):
    if get_role(request.user) != 'manager':
        return JsonResponse({'error': 'غير مصرح'}, status=403)
    _load_geo()
    govs = sorted(_GOVS, key=lambda g: g['name_ar'])
    return JsonResponse({'govs': govs})


# ══════════════════════════════════════════════════════════════════
# EXPORT — Excel & PDF (for all users)
# ══════════════════════════════════════════════════════════════════
@login_required(login_url='login')
def export_excel_view(request):
    from .exports import export_excel
    from django.http import HttpResponse
    role = get_role(request.user)
    qs = Violation.objects.select_related('governorate').all()
    if role not in ('supervisor','manager'):
        qs = qs.filter(status='approved')

    gov      = request.GET.get('gov','')
    district = request.GET.get('district','')
    village  = request.GET.get('village','')
    status   = request.GET.get('status','')
    min_area = request.GET.get('min_area','')
    search   = request.GET.get('search','')

    filters = {}
    if gov:      qs = qs.filter(governorate__pcode=gov);  filters['gov'] = gov
    if district: qs = qs.filter(district_name=district);  filters['district'] = district
    if village:  qs = qs.filter(village=village);          filters['village'] = village
    if status and role in ('supervisor','manager'):
        qs = qs.filter(status=status); filters['status'] = status
    if min_area:
        try: qs = qs.filter(area_total__gte=float(min_area)); filters['min_area'] = min_area
        except: pass
    if search:
        qs = qs.filter(Q(occupant__icontains=search)|Q(code__icontains=search)|
                       Q(village__icontains=search))
        filters['search'] = search

    records = list(qs.values(
        'code','governorate__name_ar','district_name','village','occupant',
        'basin','description','area_nile_bank','area_public','area_outside',
        'area_total','status',
    ))

    buf = export_excel(records, filters, request.user)
    log_action(request, 'export', f'Excel — {len(records)} سجل')
    response = HttpResponse(buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="violations_{len(records)}.xlsx"'
    return response


@login_required(login_url='login')
def export_pdf_view(request):
    from .exports import export_pdf
    from django.http import HttpResponse
    role = get_role(request.user)
    qs = Violation.objects.select_related('governorate').all()
    if role not in ('supervisor','manager'):
        qs = qs.filter(status='approved')

    gov      = request.GET.get('gov','')
    district = request.GET.get('district','')
    status   = request.GET.get('status','')
    min_area = request.GET.get('min_area','')
    filters  = {}
    if gov:      qs = qs.filter(governorate__pcode=gov);  filters['gov'] = gov
    if district: qs = qs.filter(district_name=district);  filters['district'] = district
    if status and role in ('supervisor','manager'):
        qs = qs.filter(status=status); filters['status'] = status
    if min_area:
        try: qs = qs.filter(area_total__gte=float(min_area)); filters['min_area'] = min_area
        except: pass

    records = list(qs.values(
        'code','governorate__name_ar','district_name','village','occupant',
        'basin','description','area_nile_bank','area_public','area_outside',
        'area_total','status',
    ))

    buf = export_pdf(records, filters, request.user)
    log_action(request, 'export', f'PDF — {len(records)} سجل')
    response = HttpResponse(buf.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="violations_{len(records)}.pdf"'
    return response


# ══════════════════════════════════════════════════════════════════
# USAGE TRACKING — تتبع الاستغلال وحساب مقابل الانتفاع
# ══════════════════════════════════════════════════════════════════
from .models import MinistryDecision, UsageType, ViolationUsage
from decimal import Decimal
from django.template.loader import render_to_string

@login_required(login_url='login')
def usage_list_api(request, pk):
    """قائمة سجلات الاستغلال لقطعة أرض + الإجمالي"""
    v = get_object_or_404(Violation, pk=pk)
    usages = ViolationUsage.objects.filter(violation=v).select_related(
        'usage_type', 'created_by', 'approved_by'
    ).order_by('date_from')

    role = get_role(request.user)
    # المشرف والمدير يرون كل السجلات — مدخل البيانات يرى المعتمدة وسجلاته
    if role not in ('supervisor', 'manager'):
        usages = usages.filter(
            models.Q(status='approved') | models.Q(created_by=request.user)
        )

    def fmt_date(d):
        return d.strftime('%Y-%m-%d') if d else None

    records = []
    for u in usages:
        records.append({
            'id':              u.id,
            'occupant_name':   u.occupant_name,
            'activity_desc':   u.activity_desc,
            'usage_type_id':   u.usage_type_id,
            'usage_type_name': u.usage_type.name if u.usage_type else '—',
            'basis':           u.basis,
            'zone':            u.zone,
            'zone_label':      u.get_zone_display(),
            'area':            u.area,
            'date_from':       fmt_date(u.date_from),
            'date_to':         fmt_date(u.date_to),
            'is_ongoing':      u.is_ongoing,
            'days':            u.days,
            'rate_per_year':   float(u.rate_per_year),
            'calculated_value':float(u.calculated_value),
            'amount_before_2021': float(u.amount_before_2021),
            'amount_paid':     float(u.amount_paid),
            'remaining':       float(u.remaining),
            'total_amount':    float(u.total_amount),
            'status':          u.status,
            'notes':           u.notes,
            'created_by':      u.created_by.get_full_name() or u.created_by.username if u.created_by else '—',
            'created_at':      u.created_at.strftime('%Y-%m-%d') if u.created_at else None,
        })

    approved = [u for u in usages if u.status == 'approved']
    total_value   = sum(float(u.calculated_value) for u in approved)
    total_before  = sum(float(u.amount_before_2021) for u in approved)
    total_paid    = sum(float(u.amount_paid)       for u in approved)
    total_remain  = (total_value + total_before) - total_paid

    return JsonResponse({
        'records':         records,
        'total_value':     total_value,
        'total_before_2021': total_before,
        'total_paid':      total_paid,
        'total_remain':    total_remain,
        'count':           len(records),
    })


@login_required(login_url='login')
def usage_add_api(request, pk):
    """إضافة سجل استغلال جديد"""
    v    = get_object_or_404(Violation, pk=pk)
    role = get_role(request.user)
    if role not in ('data_entry', 'supervisor', 'manager'):
        return JsonResponse({'error': 'غير مصرح'}, status=403)
    if request.method != 'POST':
        return JsonResponse({'error': 'POST only'}, status=405)
    try:
        data = json.loads(request.body)
    except Exception:
        return JsonResponse({'error': 'بيانات غير صالحة'}, status=400)

    required = ['occupant_name', 'date_from', 'area']
    missing  = [f for f in required if not str(data.get(f, '')).strip()]
    if missing:
        return JsonResponse({'error': f'حقول مطلوبة: {", ".join(missing)}'}, status=400)

    area = float(data.get('area', 0))
    used = ViolationUsage.objects.filter(violation=v).exclude(pk=data.get('id')).aggregate(
        s=models.Sum('area'))['s'] or 0
    if area + float(used) > v.area_total:
        return JsonResponse({'error': f'المساحة المستغلة مع الاستغلالات السابقة ({area:.2f}+{float(used):.2f}={area+float(used):.2f} م²) تتجاوز إجمالي مساحة القطعة ({v.area_total:.2f} م²)'}, status=400)

    try:
        from datetime import date as date_type
        from datetime import datetime

        date_from = datetime.strptime(data['date_from'], '%Y-%m-%d').date()
        date_to   = None
        is_ongoing= bool(data.get('is_ongoing', False))
        if not is_ongoing and data.get('date_to'):
            date_to = datetime.strptime(data['date_to'], '%Y-%m-%d').date()

        usage_type = None
        if data.get('usage_type_id'):
            usage_type = UsageType.objects.filter(pk=data['usage_type_id']).first()

        # الفئة السنوية من نوع الاستغلال + المنطقة
        zone = data.get('zone', 'other')
        rate = 0
        if usage_type:
            rate = usage_type.get_rate(zone)

        status = 'approved' if role in ('supervisor', 'manager') else 'pending'

        basis = data.get('basis', '149')
        if usage_type:
            basis = usage_type.decision

        u = ViolationUsage.objects.create(
            violation      = v,
            usage_type     = usage_type,
            basis          = basis,
            occupant_name  = data['occupant_name'].strip(),
            activity_desc  = data.get('activity_desc', '').strip(),
            zone           = zone,
            area           = area,
            date_from      = date_from,
            date_to        = date_to,
            is_ongoing     = is_ongoing,
            rate_per_year  = Decimal(str(rate)),
            amount_before_2021 = Decimal(str(data.get('amount_before_2021', 0) or 0)),
            amount_paid    = Decimal(str(data.get('amount_paid', 0) or 0)),
            notes          = data.get('notes', '').strip(),
            status         = status,
            created_by     = request.user,
        )
        # حساب القيمة تلقائياً
        u.calculated_value = Decimal(str(u.calculate_value()))
        u.save(update_fields=['calculated_value'])

        log_action(request, 'add', f'سجل استغلال — {v.code} — {u.occupant_name}')
        return JsonResponse({
            'success': True,
            'id':      u.id,
            'calculated_value': float(u.calculated_value),
            'message': 'تم الحفظ وإرساله للمراجعة' if status == 'pending' else 'تم الحفظ والاعتماد',
        })
    except Exception as e:
        import traceback
        return JsonResponse({'error': str(e), 'trace': traceback.format_exc()}, status=500)


@login_required(login_url='login')
def usage_approve_api(request, pk, usage_id):
    """موافقة/رفض سجل استغلال"""
    role = get_role(request.user)
    if role not in ('supervisor', 'manager'):
        return JsonResponse({'error': 'غير مصرح'}, status=403)
    try:
        data   = json.loads(request.body)
        action = data.get('action')  # 'approve' or 'reject'
        u = get_object_or_404(ViolationUsage, pk=usage_id, violation_id=pk)
        u.status      = 'approved' if action == 'approve' else 'rejected'
        u.approved_by = request.user
        u.save(update_fields=['status', 'approved_by'])
        log_action(request, 'approve', f'سجل استغلال #{usage_id} — {u.get_status_display()}')
        return JsonResponse({'success': True, 'status': u.status})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required(login_url='login')
def usage_calculate_preview(request, pk):
    """حساب مبدئي لمقابل الانتفاع قبل الحفظ — يستخدم سعر نوع الاستغلال حسب المنطقة"""
    try:
        data = json.loads(request.body)
        from datetime import datetime
        from django.utils import timezone

        date_from  = datetime.strptime(data['date_from'], '%Y-%m-%d').date()
        is_ongoing = bool(data.get('is_ongoing', False))
        date_to    = timezone.now().date() if is_ongoing else (
            datetime.strptime(data['date_to'], '%Y-%m-%d').date()
            if data.get('date_to') else timezone.now().date()
        )
        area  = float(data.get('area', 0))
        zone  = data.get('zone', 'other')
        usage_type_id = data.get('usage_type_id')

        # الحصول على السعر من نوع الاستغلال حسب المنطقة
        rate = 0
        if usage_type_id:
            ut = UsageType.objects.filter(pk=usage_type_id).first()
            if ut:
                rate = ut.get_rate(zone)

        if rate == 0:
            return JsonResponse({'success': True, 'breakdown': [], 'total_per_m2': 0,
                                 'total_value': 0, 'days_total': 0, 'area': area})

        days_total = (date_to - date_from).days
        total = round(days_total * (rate / 365.0) * area, 2)

        # تفصيل واحد بالقرارات المتعاقبة
        decisions = MinistryDecision.objects.filter(
            date_from__lte=date_to
        ).order_by('date_from')
        breakdown = []

        for dec in decisions:
            d_start = dec.date_from
            d_end   = dec.date_to or date_to
            overlap_start = max(date_from, d_start)
            overlap_end   = min(date_to, d_end)
            if overlap_start > overlap_end:
                continue
            days      = (overlap_end - overlap_start).days
            val_per_m2= days * (rate / 365.0)
            breakdown.append({
                'decision':    dec.decision_number,
                'date_from':   overlap_start.strftime('%Y-%m-%d'),
                'date_to':     overlap_end.strftime('%Y-%m-%d'),
                'days':        days,
                'rate':        rate,
                'value_per_m2':round(val_per_m2, 4),
            })

        return JsonResponse({
            'success':      True,
            'breakdown':    breakdown,
            'total_per_m2': round(days_total * (rate / 365.0), 4),
            'total_value':  total,
            'days_total':   days_total,
            'area':         area,
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required(login_url='login')
def usage_types_api(request):
    """قائمة أنواع الاستغلال للقوائم المنسدلة"""
    qs = UsageType.objects.filter(is_active=True)
    decision = request.GET.get('decision', '')
    if decision:
        qs = qs.filter(decision=decision)
    types = list(qs.values(
        'id', 'name', 'article', 'decision',
        'rate_warraq', 'rate_aswan', 'rate_other',
        'rate_urban_in', 'rate_urban_out',
    ))
    return JsonResponse({'types': types})


# ══════════════════════════════════════════════════════════════════
# PRINT — طباعة سجلات الاستغلال
# ══════════════════════════════════════════════════════════════════

def _calc_usage_breakdown(u):
    """حساب تفصيل مقابل الانتفاع لسجل استغلال — يستخدم سعر النوع حسب المنطقة"""
    from datetime import date as date_type
    end_date = u.date_to or date_type.today()
    start = u.date_from
    if not u.usage_type:
        return []
    rate = u.usage_type.get_rate(u.zone)
    if rate == 0:
        return []
    decisions = MinistryDecision.objects.filter(date_from__lte=end_date).order_by('date_from')
    rows = []
    for dec in decisions:
        d_start = dec.date_from
        d_end = dec.date_to or end_date
        overlap_start = max(start, d_start)
        overlap_end = min(end_date, d_end)
        if overlap_start > overlap_end:
            continue
        days = (overlap_end - overlap_start).days
        val_per_m2 = days * (rate / 365.0)
        rows.append({
            'decision': dec.decision_number,
            'date_from': overlap_start,
            'date_to': overlap_end,
            'days': days,
            'rate': rate,
            'value_per_m2': round(val_per_m2, 4),
            'value': round(val_per_m2 * u.area, 2),
        })
    return rows


@login_required(login_url='login')
def usage_print_single(request, pk, usage_id):
    """صفحة طباعة سجل استغلال واحد مع جدول الحساب"""
    from django.utils import timezone
    u = get_object_or_404(ViolationUsage, pk=usage_id, violation_id=pk)
    v = u.violation

    breakdown = _calc_usage_breakdown(u)
    total_per_m2 = sum(r['value_per_m2'] for r in breakdown)
    total_value = round(total_per_m2 * u.area, 2)

    context = {
        'usage': u,
        'violation': v,
        'breakdown': breakdown,
        'total_per_m2': round(total_per_m2, 4),
        'total_value': total_value,
        'grand_total': float(u.calculated_value) + float(u.amount_before_2021),
        'print_date': timezone.now(),
        'user': request.user,
    }
    return render(request, 'violations/usage_print_single.html', context)


@login_required(login_url='login')
def usage_print_all(request, pk):
    """صفحة طباعة كل سجلات الاستغلال لقطعة أرض مع الحسابات"""
    from django.utils import timezone
    v = get_object_or_404(Violation, pk=pk)
    usages = ViolationUsage.objects.filter(violation=v).order_by('date_from')

    usage_data = []
    for u in usages:
        breakdown = _calc_usage_breakdown(u)
        total_per_m2 = sum(r['value_per_m2'] for r in breakdown)
        usage_data.append({
            'usage': u,
            'breakdown': breakdown,
            'total_per_m2': round(total_per_m2, 4),
            'total_value': round(total_per_m2 * u.area, 2),
        })

    total_calc = sum(float(u.calculated_value) for u in usages if u.status == 'approved')
    total_before = sum(float(u.amount_before_2021) for u in usages if u.status == 'approved')

    context = {
        'violation': v,
        'usage_data': usage_data,
        'summary': {
            'count': usages.count(),
            'total_calculated': total_calc,
            'total_before_2021': total_before,
            'total_grand': total_calc + total_before,
            'total_area': v.area_total,
        },
        'print_date': timezone.now(),
        'user': request.user,
    }
    return render(request, 'violations/usage_print_all.html', context)


# ══════════════════════════════════════════════════════════════════
# 🤖 Chatbot API
# ══════════════════════════════════════════════════════════════════
from .services import chatbot_nlp


@login_required(login_url='login')
def chatbot_api(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'POST only'}, status=405)
    try:
        data = json.loads(request.body)
    except Exception:
        return JsonResponse({'error': 'بيانات غير صالحة'}, status=400)

    message = (data.get('message') or '').strip()
    if not message:
        return JsonResponse({'response': 'اكتب سؤالك', 'type': 'text'})

    parsed = chatbot_nlp.parse(message)
    if parsed['type'] == 'query':
        parsed = _execute_chatbot_query(parsed)
    return JsonResponse(parsed)


def _execute_chatbot_query(parsed: dict) -> dict:
    filters = parsed.get('filters', {})
    intent = parsed.get('intent', 'list')
    obj = parsed.get('obj', 'violation')
    limit = parsed.get('limit', 10)

    try:
        if obj == 'violation':
            return _query_violations(filters, intent, limit)
        elif obj in ('usage', 'usage_value'):
            return _query_usages(filters, intent, obj)
        else:
            return {'response': 'لم أفهم الطلب', 'data': None, 'type': 'text'}
    except Exception as e:
        return {'response': f'حدث خطأ: {str(e)}', 'data': None, 'type': 'text'}


def _build_violation_q(filters: dict) -> Q:
    q = Q()
    if filters.get('code'):
        q &= Q(code=filters['code'])
    if filters.get('status'):
        q &= Q(status=filters['status'])
    if filters.get('governorate'):
        try:
            gov = Governorate.objects.get(name_en=filters['governorate'])
            q &= Q(governorate=gov)
        except Governorate.DoesNotExist:
            pass
    # Violation-level comparison (area_total)
    if filters.get('comparison_field') == 'area' and filters.get('comparison_op'):
        op = filters['comparison_op']; val = filters['comparison_val']
        if op == 'gt': q &= Q(area_total__gt=val)
        elif op == 'lt': q &= Q(area_total__lt=val)
        elif op == 'eq': q &= Q(area_total=val)
    return q


def _build_usage_q(filters: dict) -> Q:
    q = Q()
    if filters.get('zone'):
        q &= Q(zone=filters['zone'])
    if filters.get('decision'):
        q &= Q(basis=filters['decision'])
    if filters.get('status'):
        q &= Q(status=filters['status'])
    if filters.get('code'):
        q &= Q(violation__code=filters['code'])
    if filters.get('governorate'):
        try:
            gov = Governorate.objects.get(name_en=filters['governorate'])
            q &= Q(violation__governorate=gov)
        except Governorate.DoesNotExist:
            pass
    if filters.get('comparison_field') and filters.get('comparison_op') and filters.get('comparison_val'):
        field = 'area' if filters['comparison_field'] == 'area' else 'calculated_value'
        op = filters['comparison_op']; val = filters['comparison_val']
        if op == 'gt': q &= Q(**{f'{field}__gt': val})
        elif op == 'lt': q &= Q(**{f'{field}__lt': val})
        elif op == 'eq': q &= Q(**{f'{field}': val})
    return q


def _query_violations(filters: dict, intent: str, limit: int) -> dict:
    vq = _build_violation_q(filters)
    # If usage-level filters exist, join through usages
    usage_filters = {k: v for k, v in filters.items()
                     if k in ('zone', 'decision') and v}
    if usage_filters or (filters.get('comparison_field') == 'value'):
        uq = _build_usage_q(filters)
        v_ids = ViolationUsage.objects.filter(uq).values_list('violation_id', flat=True).distinct()
        vq &= Q(id__in=v_ids)

    if intent == 'count':
        total = Violation.objects.filter(vq).count()
        return {'response': f'عدد المخالفات: {total}', 'data': {'count': total}, 'type': 'count'}

    qs = Violation.objects.filter(vq).select_related('governorate')[:limit]
    rows = []
    for v in qs:
        gov = v.governorate.name_ar if v.governorate else '—'
        rows.append({
            'id': v.id, 'code': v.code, 'occupant': v.occupant,
            'village': v.village, 'governorate': gov,
            'area_total': v.area_total, 'status': v.get_status_display(),
        })

    count = Violation.objects.filter(vq).count()
    summary = f'عرض {len(rows)} من {count} مخالفة' if count > limit else f'عدد المخالفات: {count}'
    return {'response': summary, 'data': rows, 'type': 'violation_list'}


def _query_usages(filters: dict, intent: str, obj: str) -> dict:
    from django.db.models import Sum, Avg
    uq = _build_usage_q(filters)
    qs = ViolationUsage.objects.filter(uq)

    if intent == 'count':
        total = qs.count()
        return {'response': f'عدد سجلات الاستغلال: {total}', 'data': {'count': total}, 'type': 'count'}

    if intent == 'sum' or obj == 'usage_value':
        agg = qs.aggregate(total=Sum('calculated_value'), total_paid=Sum('amount_paid'))
        total = float(agg['total'] or 0)
        paid = float(agg['total_paid'] or 0)
        return {
            'response': f'إجمالي مقابل الانتفاع: {total:,.2f} جنيه — المسدد: {paid:,.2f} جنيه — المتبقي: {total - paid:,.2f} جنيه',
            'data': {'total': total, 'paid': paid, 'remaining': total - paid},
            'type': 'sum',
        }

    if intent == 'avg':
        agg = qs.aggregate(avg=Avg('calculated_value'))
        avg_val = float(agg['avg'] or 0)
        return {'response': f'متوسط قيمة الاستغلال: {avg_val:,.2f} جنيه', 'data': {'avg': avg_val}, 'type': 'avg'}

    rows = []
    for u in qs.select_related('violation', 'usage_type')[:20]:
        rows.append({
            'id': u.id, 'violation_code': u.violation.code,
            'occupant_name': u.occupant_name,
            'usage_type': u.usage_type.name if u.usage_type else '—',
            'zone': u.get_zone_display(), 'area': u.area,
            'calculated_value': float(u.calculated_value),
            'status': u.get_status_display(),
        })

    count = qs.count()
    summary = f'عرض {len(rows)} من {count} سجل استغلال' if count > 20 else f'عدد السجلات: {count}'
    return {'response': summary, 'data': rows, 'type': 'usage_list'}


# ══════════════════════════════════════════════════════════════════
# 🛰️ Satellite imagery API
# ══════════════════════════════════════════════════════════════════
from .services import satellite as satellite_svc


@login_required(login_url='login')
def satellite_search_api(request):
    lat = request.GET.get('lat')
    lng = request.GET.get('lng')
    if not lat or not lng:
        return JsonResponse({'error': 'lat,lng مطلوبان'}, status=400)

    try:
        lat, lng = float(lat), float(lng)
    except ValueError:
        return JsonResponse({'error': 'lat,lng غير صحيحين'}, status=400)

    date_from = request.GET.get('date_from') or '2024-06-01'
    date_to   = request.GET.get('date_to')   or '2024-12-31'

    result = satellite_svc.search_best(lat, lng, date_from, date_to)
    if not result:
        return JsonResponse({'error': 'لا توجد صور مناسبة لهذه الفترة'}, status=404)
    result['bounds'] = [lng - 0.01, lat - 0.01, lng + 0.01, lat + 0.01]
    return JsonResponse(result)


@login_required(login_url='login')
def satellite_compare_api(request):
    lat = request.GET.get('lat')
    lng = request.GET.get('lng')
    if not lat or not lng:
        return JsonResponse({'error': 'lat,lng مطلوبان'}, status=400)
    try:
        lat, lng = float(lat), float(lng)
    except ValueError:
        return JsonResponse({'error': 'lat,lng غير صحيحين'}, status=400)

    year_before = int(request.GET.get('year_before', 2022))
    year_after  = int(request.GET.get('year_after', 2024))

    result = satellite_svc.search_before_after(lat, lng, year_before, year_after)
    bounds = [lng - 0.01, lat - 0.01, lng + 0.01, lat + 0.01]
    if result.get('before'): result['before']['bounds'] = bounds
    if result.get('after'):  result['after']['bounds']  = bounds
    return JsonResponse(result)
